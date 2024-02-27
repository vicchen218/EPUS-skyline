import os, sys
sys.path.append(os.path.abspath(os.pardir))
from copy import copy,deepcopy
import pandas as pd
import pickle
import time
import openpyxl
from data.dataClass import Data, batchImport
from skyline.slideUPSky import slideUPSky
from skyline.PSky import PSky
# from visualize import visualize
from visualize.visualize import visualize

# PSky class use only in server side 
class servePSky(PSky):
    def __init__(self, dim, ps, radius, drange=[0,1000], wsize=10):
        """
        Initializer
        :param dim: int
            The dimension of data
        :param ps: int
            The occurance count of the instance.
        :param radius: int
            radius use to prevent data being pruned unexpectedly.
            Recommand to be set according to the name of .csv file.
        :param drange: list(int)
            data range [min, max]
        :param wsize: int
            Size of sliding window.
            Note that the window size should be sum(edge window)
        """
        PSky.__init__(self, dim, ps, radius, drange, wsize)
    def receive(self,data):
        """
        Update data received by server
        :param data: dict
            json format(dict) data include change of an edge node.
            Delete: outdated data
            SK1: new data in skyline set
            SK2: new data in skyline2 set
        """
        # print("data list is",data)
        # print("data[0] list is",data[0])
        if len(data[0]['Delete']) > 0:
            # print("get in delete")
            for d in data[0]['Delete']:
                if d in self.window:
                    self.window.remove(d)
                    self.outdated.append(d)
                    self.updateIndex(d, 'remove')
        if len(data[0]['SK1']) > 0:
            # print("get in sk1")
            for d in data[0]['SK1']:
                if d not in self.window:
                    self.window.append(d)
                    self.skyline.append(d)
                    self.updateIndex(d, 'insert')
                elif d in self.skyline2:
                    self.skyline2.remove(d)
                    self.skyline.append(d)
                # ignore other condition
        if len(data[0]['SK2']) > 0:
            # print("get in sk2")
            for d in data[0]['SK2']:
                if d not in self.window:
                    self.window.append(d)
                    self.skyline2.append(d)
                    self.updateIndex(d, 'insert')
                elif d in self.skyline:
                    self.skyline.remove(d)
                    self.skyline2.append(d)
                # ignore other condition
        # self.update()
    def update(self):
        """
        Update global skyline set
        """
        if len(self.outdated) > 0:
            # Remove outdated data in sk2
            for d in self.outdated:
                if d in self.skyline2:
                    self.skyline2.remove(d)
            # Remove outdated data in sk, add sk2 data to sk when needed
            for d in self.outdated:
                if d in self.skyline:
                    self.skyline.remove(d)
                    sstart = [ i for i in d.getLocationMax()]
                    send = [self.drange[1] for i in range(self.dim)]
                    search = [ p.object for p in (self.index.intersection(tuple(sstart+send),objects=True))]
                    for sd in search:
                        if sd in self.skyline2:
                            self.skyline2.remove(sd)
                            self.skyline.append(sd)
            # Clear outdated data
            self.outdated = []
        # prune objects in sk, move data dominated by other sk point to sk2
        for d in self.skyline.copy():
            if d in self.skyline:
                vurstart = [ self.drange[1] if i+2*self.radius+0.1 > self.drange[1] else i+2*self.radius+0.1 for i in d.getLocationMax()]
                vurend = [ self.drange[1] for i in range(self.dim)]
                vur = [ p.object for p in (self.index.intersection(tuple(vurstart+vurend),objects=True))]
                for p in vur:
                    if p in self.skyline:
                        self.skyline.remove(p)
                        self.skyline2.append(p)
        # prune objects in sk2
        for d in self.skyline2.copy():
            if d in self.skyline2:
                vurstart = [ self.drange[1] if i+2*self.radius+0.1 > self.drange[1] else i+2*self.radius+0.1 for i in d.getLocationMax()]
                vurend = [ self.drange[1] for i in range(self.dim)]
                vur = [ p.object for p in (self.index.intersection(tuple(vurstart+vurend),objects=True))]
                for p in vur:
                    if p in self.skyline2:
                       self.skyline2.remove(p)

if __name__ == "__main__":
    rank = (2, 4, 6,8,10,12,14,16)
    dimlist = (3, 4, 5)
    rowname = ('Dimension','Node Number', 'Mean Edge Processing Time', 'Max Edge Processing Time', 'Server Processing Time',
            'Process Latency', 'Transmission Latency with 1Mbps', 'Transmission Latency with 200Kbps','Total Transmission')

    

    # for it in range(10):
    # 創建空的 DataFrame
    df = pd.DataFrame(columns=rowname)
    for i in range(5):
        # df = pd.DataFrame(columns=rowname)
        for dim in dimlist:
            for num in rank:
                edgenum = num
                etmax = []
                row_index = len(df)
                # row = 15 * (dim - 3) + (num / 2) + 2
                print("----- amount of nodes ", edgenum, " ------")
                df.loc[row_index, 'Dimension'] = dim
                df.loc[row_index, 'Node Number'] = num
                print("dim is", dim)
                for k in range(edgenum):
                    # 執行其他運算
                    eid = str(k)
                    usky = slideUPSky(dim, 3, 3, [0, 100], wsize=10)
                    dqueue = batchImport('20_dim' + str(dim) + '_pos3_rad3_0100.csv', 3)
                    idx = [i for i in range(20) if i % edgenum == k]
                    
                    with open('pickle_edge' + eid + '.pickle', 'wb') as f:
                        start_time = time.time()
                        for i in idx:
                            oldsk = usky.getSkyline().copy()
                            oldsk2 = usky.getSkyline2().copy()
                            usky.receiveData(dqueue[i])
                            out = usky.getOutdated().copy()
                            usky.updateSkyline()
                            usk1 = list(set(usky.getSkyline()) - set(oldsk))
                            usk2 = list(set(usky.getSkyline2()) - set(oldsk2))
                            result = {'Delete': out, 'SK1': usk1, 'SK2': usk2}
                            pickle.dump(result, f)
                        finish_time = time.time() - start_time
                        etmax.append(finish_time)
                        print("edge", k+1, "process --- %s seconds ---" % (finish_time))
                            
                    usky.removeRtree()
                    
                edgemean = sum(etmax) / len(etmax)
                edgemax = max(etmax)

                # 將運算結果存入指定的欄位
                df.loc[row_index, 'Mean Edge Processing Time'] = edgemean
                df.loc[row_index, 'Max Edge Processing Time'] = edgemax




                ### template_picklefile

                edgedata =[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
                templist =[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
                for e in range(edgenum):
                    idx = [i for i in range(20) if i%edgenum == e]
                    with open('pickle_edge'+str(e)+'.pickle', 'rb') as f:
                        for d in idx:
                            edgedata[e].append(pickle.load(f))

                templist=deepcopy(edgedata) #for wsize test


                # 刪除 pickle 檔案
                for e in range(edgenum):
                    file_path = 'pickle_edge' + str(e) + '.pickle'
                    if os.path.exists(file_path):
                        os.remove(file_path)


                ###catch communication load
                sdatalist =[]
                # r.write('-- Transmission cost with dim {a}--\n'.format(a=dim))
                for k in range(edgenum):
                    # print("\n\n")
                    sdata =0
                    for m in range(len(templist[k])):
                        stemp = len(templist[k][m]['Delete'])+len(templist[k][m]['SK1'])+len(templist[k][m]['SK2'])
                        sdata =sdata + stemp
                    print("node ", k, "send", sdata) 
                    # r.write('node {a} send {b}\n'.format(a=k,b=sdata))
                    sdatalist.append(sdata)
                print("total transmission", sum(sdatalist))
                # r.write('total transmission {a}\n\n'.format(a=sum(sdatalist) ))
                df.loc[row_index, 'Total Transmission']=sum(sdatalist)
                df.loc[row_index,'Transmission Latency with 1Mbps']=df.loc[row_index,'Total Transmission']*0.003/(1*num)
                df.loc[row_index,'Transmission Latency with 200Kbps']=df.loc[row_index,'Total Transmission']*0.003/(0.2*num)
                
                
                
                ###localserver
                # for sw in (100,300,500,700):#for wsize test
                # defult sw=ew*edge_num
                sw = 10*edgenum
                skyServer = servePSky(dim, 3, 3, [0,100], wsize=sw)
                server_time = time.time()-time.time() # let time be 0
                for k in range(20):
                    # pop list node by node
                    m = k % edgenum # node by node            
                    start_time = time.time()
                    skyServer.receive(edgedata[m])
                    skyServer.update()
                    t=time.time() - start_time # just calculate the recieve and update time
                    server_time = server_time+t
                    edgedata[m].pop(0)
                print("server-windowsize is",sw)
                print("--- finish --- %s seconds ---" % (server_time))
                df.loc[row_index, 'Server Processing Time']=server_time
                skyServer.removeRtree()
                edgedata =[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]#for wsize test
                edgedata=deepcopy(templist)#for wsize test 
                # s1.cell(row,5).value=s1.cell(row,4).value+s1.cell(row,3).value               


        # 將Process latency 做組合
        # 將 DataFrame 寫入 Excel 檔案
        df['Process Latency']=df['Max Edge Processing Time']+df['Server Processing Time']
        # book = openpyxl.load_workbook('Distributed_Dimension_test.xlsx')
        # writer = pd.ExcelWriter('Distributed_Dimension_test.xlsx', engine='xlsxwriter') 
        # writer.book = book
        # df.to_excel(writer, sheet_name='Test'+str(i), index=False)
        # df.to_excel(writer, sheet_name='Test' + str(i), index_label=False)
        
        df.to_excel('Distributed_Dimension_test.xlsx', sheet_name='Test', index_label=False)
        # writer.save()
        # writer.close()
        # df.drop(df.index,inplace=True)